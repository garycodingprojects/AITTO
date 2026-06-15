"""XML input/output helpers for the aiTTO GA toolbox."""

from __future__ import annotations

import json
import xml.etree.ElementTree as ET
from pathlib import Path
from xml.dom import minidom

from openpyxl import load_workbook

from .model import (
    ClassAssignment,
    ClassSection,
    ConstraintDefinition,
    Course,
    DEFAULT_HARD_CONSTRAINTS,
    DEFAULT_DATASET_XML,
    DEFAULT_SOFT_CONSTRAINTS,
    DistributionConstraint,
    ProblemInstance,
    Room,
    RoomOption,
    Student,
    TimeOption,
    TimetableSolution,
    UnavailablePeriod,
)


def _int_attr(element: ET.Element, name: str, default: int = 0) -> int:
    """Read an integer XML attribute with a safe default."""

    return int(element.get(name, str(default)))


def _constraint_workbook_path(source_path: Path, filename: str) -> Path:
    """Locate a constraint workbook beside the dataset folder when possible."""

    dataset_dir = source_path.parent.parent if source_path.parent.name == "itc2019" else source_path.parent
    return dataset_dir / filename


def _constraint_support(category: str, constraint_type: str, description: str) -> tuple[bool, str]:
    """Map workbook constraint rows onto prototype evaluator behaviours."""

    text = f"{constraint_type} {description}".lower()
    if category == "hard":
        if "clash" in text:
            return True, "same-cohort overlaps are counted as hard cohort_overlap violations"
        if "missing allocation" in text:
            return True, "required student/course allocations are checked against active cohort classes"
        if "extra allocation" in text:
            return True, "more than one active class per student/course is counted as hard extra_allocation"
        if "incorrect allocation" in text:
            return True, "classes assigned to students who do not take the course are counted as hard incorrect_allocation"
        if "one course" in text and "room" in text:
            return True, "room double-booking is counted as hard room_overlap"
        if "capacity" in text:
            return True, "class limit above room capacity is counted as hard room_capacity"
    if category == "soft":
        if "day preference" in text:
            return True, "compiled XML time penalties approximate day preferences"
        if "time period preference" in text:
            return True, "compiled XML time penalties approximate time-period preferences"
        if "single course" in text:
            return True, "exactly one cohort class meeting on a day is penalized as single_course_day_penalty"
    return False, "not enough structured data is available in the current prototype"


def load_constraint_definitions(path: str | Path, category: str) -> tuple[ConstraintDefinition, ...]:
    """Load hard/soft constraint descriptions from a simple workbook catalog."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        return ()

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    definitions: list[ConstraintDefinition] = []
    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().lower() for value in next(rows, ())]
        if "type" not in headers or "description" not in headers:
            continue
        type_index = headers.index("type")
        description_index = headers.index("description")
        for row in rows:
            constraint_type = str(row[type_index] or "").strip()
            description = str(row[description_index] or "").strip()
            if not constraint_type or not description:
                continue
            supported, implementation = _constraint_support(category, constraint_type, description)
            definitions.append(
                ConstraintDefinition(
                    category=category,
                    constraint_type=constraint_type,
                    description=description,
                    supported=supported,
                    implementation=implementation,
                )
            )
    return tuple(definitions)


def load_problem(path: str | Path = DEFAULT_DATASET_XML) -> ProblemInstance:
    """Parse an aiTTO XML dataset into a compact `ProblemInstance`."""

    source_path = Path(path)
    root = ET.parse(source_path).getroot()

    weights = {"time": 1, "room": 1, "distribution": 1, "student": 1}
    optimization = root.find("./optimization")
    if optimization is not None:
        weights.update({key: _int_attr(optimization, key, 1) for key in weights})

    rooms: dict[int, Room] = {}
    for room_el in root.findall("./rooms/room"):
        room_id = _int_attr(room_el, "id")
        unavailable = tuple(
            UnavailablePeriod(
                days=block.get("days", ""),
                start=_int_attr(block, "start"),
                length=_int_attr(block, "length"),
                weeks=block.get("weeks", ""),
            )
            for block in room_el.findall("./unavailable")
        )
        rooms[room_id] = Room(
            id=room_id,
            name=room_el.get("name", f"Room {room_id}"),
            capacity=_int_attr(room_el, "capacity"),
            room_type=room_el.get("type", ""),
            unavailable=unavailable,
        )

    courses: dict[int, Course] = {}
    classes: dict[int, ClassSection] = {}
    for course_el in root.findall("./courses/course"):
        course_id = _int_attr(course_el, "id")
        course_code = course_el.get("code", str(course_id))
        class_ids: list[int] = []

        for class_el in course_el.findall(".//class"):
            class_id = _int_attr(class_el, "id")
            class_ids.append(class_id)
            room_options = tuple(
                RoomOption(room_id=_int_attr(room_el, "id"), penalty=_int_attr(room_el, "penalty"))
                for room_el in class_el.findall("./room")
            )
            time_options = tuple(
                TimeOption(
                    days=time_el.get("days", ""),
                    start=_int_attr(time_el, "start"),
                    length=_int_attr(time_el, "length"),
                    weeks=time_el.get("weeks", ""),
                    penalty=_int_attr(time_el, "penalty"),
                )
                for time_el in class_el.findall("./time")
            )
            classes[class_id] = ClassSection(
                id=class_id,
                course_id=course_id,
                course_code=course_code,
                name=class_el.get("name", f"Class {class_id}"),
                cohort=class_el.get("cohort", ""),
                limit=_int_attr(class_el, "limit"),
                rooms=room_options,
                times=time_options,
            )

        courses[course_id] = Course(
            id=course_id,
            code=course_code,
            title=course_el.get("title", course_code),
            class_ids=tuple(class_ids),
        )

    students: dict[int, Student] = {}
    for student_el in root.findall("./students/student"):
        student_id = _int_attr(student_el, "id")
        students[student_id] = Student(
            id=student_id,
            external_id=student_el.get("externalId", str(student_id)),
            course_ids=tuple(_int_attr(course_el, "id") for course_el in student_el.findall("./course")),
        )

    distributions = tuple(
        DistributionConstraint(
            constraint_type=distribution_el.get("type", ""),
            penalty=_int_attr(distribution_el, "penalty"),
            required=distribution_el.get("required", "false").lower() == "true",
            class_ids=tuple(_int_attr(class_el, "id") for class_el in distribution_el.findall("./class")),
        )
        for distribution_el in root.findall("./distributions/distribution")
    )
    hard_path = _constraint_workbook_path(source_path, DEFAULT_HARD_CONSTRAINTS.name)
    soft_path = _constraint_workbook_path(source_path, DEFAULT_SOFT_CONSTRAINTS.name)

    return ProblemInstance(
        name=root.get("name", source_path.stem),
        nr_days=_int_attr(root, "nrDays"),
        nr_weeks=_int_attr(root, "nrWeeks"),
        slots_per_day=_int_attr(root, "slotsPerDay"),
        weights=weights,
        rooms=rooms,
        courses=courses,
        classes=classes,
        students=students,
        distributions=distributions,
        hard_constraints=load_constraint_definitions(hard_path, "hard"),
        soft_constraints=load_constraint_definitions(soft_path, "soft"),
        source_path=source_path,
    )


def solution_to_dict(problem: ProblemInstance, solution: TimetableSolution) -> dict[str, object]:
    """Convert a solution to a JSON-friendly dictionary."""

    classes: list[dict[str, object]] = []
    for class_id, assignment in sorted(solution.assignments.items()):
        section = problem.classes[class_id]
        time = section.times[assignment.time_index]
        room = problem.rooms.get(assignment.room_id)
        classes.append(
            {
                "class_id": class_id,
                "class_name": section.name,
                "course_code": section.course_code,
                "cohort": section.cohort,
                "room_id": assignment.room_id,
                "room_name": room.name if room else str(assignment.room_id),
                "days": time.days,
                "start": time.start,
                "length": time.length,
                "weeks": time.weeks,
            }
        )

    return {
        "problem": problem.name,
        "source": str(problem.source_path) if problem.source_path else None,
        "hard_violations": solution.hard_violations,
        "total_cost": solution.total_cost,
        "breakdown": solution.breakdown,
        "classes": classes,
    }


def write_solution_json(problem: ProblemInstance, solution: TimetableSolution, path: str | Path) -> None:
    """Write a solution as readable JSON for lightweight downstream tooling."""

    Path(path).write_text(
        json.dumps(solution_to_dict(problem, solution), indent=2),
        encoding="utf-8",
    )


def write_solution_xml(problem: ProblemInstance, solution: TimetableSolution, path: str | Path) -> None:
    """Write a compact ITC-style solution XML file for the prototype solver."""

    root = ET.Element(
        "solution",
        {
            "name": problem.name,
            "runtime": "0",
            "technique": "aiTTO Python GA prototype",
            "author": "aiTTO",
            "institution": "Prototype",
        },
    )

    for class_id, assignment in sorted(solution.assignments.items()):
        section = problem.classes[class_id]
        time = section.times[assignment.time_index]
        class_el = ET.SubElement(
            root,
            "class",
            {
                "id": str(class_id),
                "days": time.days,
                "start": str(time.start),
                "length": str(time.length),
                "weeks": time.weeks,
                "room": str(assignment.room_id),
            },
        )
        # The prototype dataset has one section per cohort/module. Assigning all
        # matching students keeps export simple until full student sectioning is added.
        for student in problem.students.values():
            if section.course_id in student.course_ids and section.cohort in student.external_id:
                ET.SubElement(class_el, "student", {"id": str(student.id)})

    rough_xml = ET.tostring(root, encoding="utf-8")
    pretty_xml = minidom.parseString(rough_xml).toprettyxml(indent="  ", encoding="utf-8")
    Path(path).write_text(pretty_xml.decode("utf-8"), encoding="utf-8")

