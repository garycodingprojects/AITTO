"""
Convert the local Excel dataset into an ITC 2019-style problem XML file.

The source workbooks in ``dataset/`` are planning spreadsheets, not a complete
ITC 2019 instance. In particular, they contain cohort/class rows rather than
individual student records, and they do not define exact weekly teaching slots.
This converter therefore creates a standards-shaped ITC 2019 instance using
documented assumptions:

* Each row in ``students.xlsx`` is treated as a cohort.
* Each cohort is expanded into synthetic individual students.
* Each module becomes one ITC course with one config and one subpart.
* Each module/cohort pair becomes one class section.
* Possible times are generated from regular weekday teaching slots.
* Holidays and non-teaching days from the academic calendar are emitted as
  room unavailability entries.

The generated XML is intended as a practical starting instance for solver
development. When real enrollment counts, teacher availability, or exact
meeting patterns become available, replace the assumptions in this script with
those source fields.
"""

from __future__ import annotations

import argparse
import math
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from xml.dom import minidom


REPO_ROOT = Path(__file__).resolve().parent.parent
DATASET_DIR = REPO_ROOT / "dataset"
DEFAULT_OUTPUT = DATASET_DIR / "itc2019" / "aitto_dataset.xml"

# aiTTO uses 30-minute slots to keep the scheduling search space manageable.
# Times are still measured from midnight, so 08:30 is slot 17.
SLOT_MINUTES = 30
DEFAULT_DAY_COUNT = 7
SLOTS_PER_DAY = 24 * 60 // SLOT_MINUTES

# Keep the generated instance compact while still giving the solver choices.
DEFAULT_COHORT_SIZE = 30
DEFAULT_SESSION_HOURS = 2
DEFAULT_TEACHING_START_TIME = "08:30"
DEFAULT_TEACHING_END_TIME = "17:30"
DEFAULT_DURATION_INCREMENT_MINUTES = 30

# Weekday indexes follow ITC binary strings: Monday, Tuesday, ..., Sunday.
WEEKDAY_INDEX = {
    "Monday": 0,
    "Tuesday": 1,
    "Wednesday": 2,
    "Thursday": 3,
    "Friday": 4,
    "Saturday": 5,
    "Sunday": 6,
}

# Common day patterns used to approximate one, two, or three meetings per week.
DAY_PATTERNS = {
    1: ["1000000", "0100000", "0010000", "0001000", "0000100"],
    2: ["1010000", "0101000", "0010100", "1001000", "0100100"],
    3: ["1010100", "0101010", "1110000"],
}

# GUI/CLI sentinel: cohort takes no modules from any module-set workbook.
NONE_MODULE_SET = "None"


@dataclass(frozen=True)
class RoomRow:
    """Source room row plus the numeric id required by ITC XML."""

    id: int
    name: str
    capacity: int
    room_type: str


@dataclass(frozen=True)
class ModuleRow:
    """A module/course row from ``modules_set_A.xlsx``."""

    id: int
    code: str
    title: str
    contact_hours: int
    room_type: str
    semesters: tuple[int, ...]


@dataclass(frozen=True)
class CohortRow:
    """A programme/class cohort from ``students.xlsx``."""

    id: int
    course: str
    class_name: str


@dataclass(frozen=True)
class ModuleSet:
    """A named group of modules that can be assigned to one or more cohorts."""

    label: str
    path: Path
    modules: tuple[ModuleRow, ...]


@dataclass(frozen=True)
class TimeSlotSettings:
    """User-editable teaching-day bounds used when generating time options."""

    start_time: str = DEFAULT_TEACHING_START_TIME
    end_time: str = DEFAULT_TEACHING_END_TIME
    duration_increment_minutes: int = DEFAULT_DURATION_INCREMENT_MINUTES


@dataclass(frozen=True)
class ConversionStats:
    """Counts reported after a conversion run."""

    output: Path
    rooms: int
    modules: int
    cohorts: int
    classes: int
    synthetic_students: int


def read_first_sheet(path: Path) -> list[tuple]:
    """Return all rows from the first worksheet, skipping temporary Excel files."""

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    return list(ws.iter_rows(values_only=True))


def normalize_header(value: object) -> str:
    """Normalize a spreadsheet header for robust column lookup."""

    return str(value or "").strip().lower()


def safe_int(value: object, default: int = 0) -> int:
    """Convert a spreadsheet value to int while tolerating blank cells."""

    if value is None or value == "":
        return default
    return int(float(value))


def load_rooms(path: Path) -> list[RoomRow]:
    """Load classroom metadata and assign sequential numeric ITC ids."""

    rows = read_first_sheet(path)
    headers = [normalize_header(v) for v in rows[0]]
    idx_name = headers.index("classroom")
    idx_capacity = headers.index("capacity")
    idx_type = headers.index("type")

    rooms: list[RoomRow] = []
    for room_id, row in enumerate(rows[1:], start=1):
        if not row[idx_name]:
            continue
        rooms.append(
            RoomRow(
                id=len(rooms) + 1,
                name=str(row[idx_name]).strip(),
                capacity=safe_int(row[idx_capacity], DEFAULT_COHORT_SIZE),
                room_type=str(row[idx_type] or "Classroom").strip(),
            )
        )
    return rooms


def load_modules(path: Path) -> list[ModuleRow]:
    """Load module rows, ignoring category headings in the workbook."""

    rows = read_first_sheet(path)
    header_row_index = next(
        i for i, row in enumerate(rows) if "Module Code" in [str(v) for v in row if v]
    )
    headers = [normalize_header(v) for v in rows[header_row_index]]

    idx_code = headers.index("module code")
    idx_title = headers.index("module title")
    idx_hours = headers.index("contact hours")
    idx_room_type = headers.index("room type")
    semester_columns = {
        int(re.search(r"\d+", header).group(0)): i
        for i, header in enumerate(headers)
        if header.startswith("sem ") and re.search(r"\d+", header)
    }

    modules: list[ModuleRow] = []
    for row in rows[header_row_index + 1 :]:
        code = row[idx_code]
        title = row[idx_title]
        if not code or not title or not str(code).strip()[0].isalnum():
            continue

        semesters = tuple(
            sem for sem, col in sorted(semester_columns.items()) if str(row[col] or "").strip().upper() == "P"
        )
        if not semesters:
            continue

        modules.append(
            ModuleRow(
                id=len(modules) + 1,
                code=str(code).strip(),
                title=str(title).strip(),
                contact_hours=safe_int(row[idx_hours], DEFAULT_SESSION_HOURS),
                room_type=str(row[idx_room_type] or "Classroom").strip(),
                semesters=semesters,
            )
        )
    return modules


def discover_module_files(dataset_dir: Path) -> list[Path]:
    """Find module-set workbooks in a dataset folder."""

    files = sorted(
        path
        for path in dataset_dir.glob("modules_set_*.xlsx")
        if not path.name.startswith("~$")
    )
    default_file = dataset_dir / "modules_set_A.xlsx"
    if not files and default_file.exists():
        files = [default_file]
    return files


def load_module_sets(paths: Iterable[Path]) -> tuple[list[ModuleRow], dict[str, ModuleSet]]:
    """
    Load module-set files and normalize duplicate module codes.

    Multiple cohorts can use different module-set files. ITC course ids must be
    globally unique, so duplicate module codes across module sets are collapsed
    into a single course id.
    """

    raw_sets: list[tuple[str, Path, list[ModuleRow]]] = []
    unique_by_code: dict[str, ModuleRow] = {}

    for path in paths:
        modules = load_modules(path)
        label = path.stem
        raw_sets.append((label, path, modules))
        for module in modules:
            unique_by_code.setdefault(module.code, module)

    normalized_by_code: dict[str, ModuleRow] = {}
    all_modules: list[ModuleRow] = []
    for module in unique_by_code.values():
        normalized = ModuleRow(
            id=len(all_modules) + 1,
            code=module.code,
            title=module.title,
            contact_hours=module.contact_hours,
            room_type=module.room_type,
            semesters=module.semesters,
        )
        normalized_by_code[module.code] = normalized
        all_modules.append(normalized)

    module_sets: dict[str, ModuleSet] = {}
    for label, path, raw_modules in raw_sets:
        module_sets[label] = ModuleSet(
            label=label,
            path=path,
            modules=tuple(normalized_by_code[module.code] for module in raw_modules),
        )

    return all_modules, module_sets


def load_cohorts(path: Path) -> list[CohortRow]:
    """Load cohort rows from the student workbook."""

    rows = read_first_sheet(path)
    headers = [normalize_header(v) for v in rows[0]]
    idx_course = headers.index("course")
    idx_class = headers.index("class")

    cohorts: list[CohortRow] = []
    for row in rows[1:]:
        if not row[idx_course] or not row[idx_class]:
            continue
        cohorts.append(
            CohortRow(
                id=len(cohorts) + 1,
                course=str(row[idx_course]).strip(),
                class_name=str(row[idx_class]).strip(),
            )
        )
    return cohorts


def load_calendar(path: Path) -> tuple[int, dict[int, set[int]], list[tuple[int, int]]]:
    """
    Read the daily academic calendar.

    Returns:
        nr_weeks: Number of academic weeks in the source calendar.
        regular_weekdays_by_week: week -> set(weekday indexes that are regular teaching days).
        unavailable_days: list of (week_index, weekday_index) pairs for non-regular days.
    """

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Daily Calendar"]
    headers = [normalize_header(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx_week = headers.index("week")
    idx_day = headers.index("day of week")
    idx_event = headers.index("event type")

    regular_weekdays_by_week: dict[int, set[int]] = {}
    unavailable_days: list[tuple[int, int]] = []
    max_week = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        week = safe_int(row[idx_week], 1)
        day_name = str(row[idx_day] or "").strip()
        event_type = str(row[idx_event] or "Regular").strip()
        if day_name not in WEEKDAY_INDEX:
            continue

        weekday = WEEKDAY_INDEX[day_name]
        max_week = max(max_week, week)

        if event_type == "Regular" and weekday < 5:
            regular_weekdays_by_week.setdefault(week, set()).add(weekday)
        else:
            unavailable_days.append((week, weekday))

    return max_week, regular_weekdays_by_week, unavailable_days


def semester_weeks(
    regular_weekdays_by_week: dict[int, set[int]],
    semesters: Iterable[int],
) -> str:
    """
    Build an ITC weeks bitstring for the given semesters.

    The source workbook marks modules as Sem 1 / Sem 2 / Sem 3, but does not
    define exact teaching weeks per semester. The ranges below follow the
    academic calendar events:

    * Semester 1: weeks 1-19, before the January revision/exam period.
    * Semester 2: weeks 22-42, before the late-June revision/exam period.
    * Semester 3: weeks 43-53, retained for future summer/third-semester data.
    """

    semester_ranges = {
        1: range(1, 20),
        2: range(22, 43),
        3: range(43, 54),
    }
    selected_weeks: set[int] = set()
    for semester in semesters:
        selected_weeks.update(semester_ranges.get(semester, []))

    if not selected_weeks:
        selected_weeks.update(regular_weekdays_by_week.keys())

    nr_weeks = max(max(regular_weekdays_by_week.keys(), default=1), 53)
    return "".join(
        "1" if week in selected_weeks and regular_weekdays_by_week.get(week) else "0"
        for week in range(1, nr_weeks + 1)
    )


def time_to_minutes(value: str) -> int:
    """Convert ``HH:MM`` into minutes after midnight for validation and slot math."""

    if not re.fullmatch(r"\d{1,2}:\d{2}", value.strip()):
        raise ValueError(f"Invalid time '{value}'. Use HH:MM, for example 08:30.")
    hour, minute = [int(part) for part in value.strip().split(":")]
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        raise ValueError(f"Invalid time '{value}'. Use a 24-hour clock value between 00:00 and 23:59.")
    return hour * 60 + minute


def minutes_to_slot(minutes: int) -> int:
    """Convert minutes after midnight into the aiTTO 30-minute slot index."""

    if minutes % SLOT_MINUTES:
        raise ValueError(f"Time value {minutes} minutes is not aligned to {SLOT_MINUTES}-minute slots.")
    return minutes // SLOT_MINUTES


def time_to_slot(value: str) -> int:
    """Convert ``HH:MM`` into an aiTTO 30-minute slot index."""

    return minutes_to_slot(time_to_minutes(value))


def validate_time_slot_settings(settings: TimeSlotSettings) -> tuple[int, int, int]:
    """
    Validate user-facing timeslot settings and return minute-based values.

    The generated XML uses aiTTO's 30-minute slots rather than ITC's original
    5-minute slots, reducing the number of possible start/length values the
    solver needs to consider.
    """

    start_minutes = time_to_minutes(settings.start_time)
    end_minutes = time_to_minutes(settings.end_time)
    increment = settings.duration_increment_minutes
    if start_minutes >= end_minutes:
        raise ValueError("Teaching start time must be earlier than teaching end time.")
    if increment <= 0:
        raise ValueError("Duration increment must be a positive number of minutes.")
    if increment % DEFAULT_DURATION_INCREMENT_MINUTES:
        raise ValueError("Duration increment must be a multiple of 30 minutes.")
    if increment % SLOT_MINUTES:
        raise ValueError(f"Duration increment must align to {SLOT_MINUTES}-minute slots.")
    if start_minutes % SLOT_MINUTES or end_minutes % SLOT_MINUTES:
        raise ValueError(f"Teaching start and end times must align to {SLOT_MINUTES}-minute slots.")
    return start_minutes, end_minutes, increment


def module_meetings_per_week(module: ModuleRow, duration_increment_minutes: int) -> tuple[int, int]:
    """
    Approximate weekly meeting count and minute length from contact hours.

    The source gives total contact hours only. We cap each meeting at roughly
    two hours by default and spread larger modules over multiple weekly days.
    The returned duration is rounded up to the configured increment so module
    lengths remain multiples of 30 minutes by default.
    """

    active_semesters = max(len(module.semesters), 1)
    nominal_weeks = 18 * active_semesters
    weekly_hours = max(module.contact_hours / nominal_weeks, 1.0)
    meetings = min(max(math.ceil(weekly_hours / DEFAULT_SESSION_HOURS), 1), 3)
    meeting_minutes = max(weekly_hours * 60 / meetings, duration_increment_minutes)
    rounded_minutes = math.ceil(meeting_minutes / duration_increment_minutes) * duration_increment_minutes
    return meetings, rounded_minutes


def make_time_options(module: ModuleRow, weeks: str, settings: TimeSlotSettings) -> list[dict[str, str | int]]:
    """Create a small domain of possible ITC time assignments for a module."""

    start_minutes, end_minutes, increment = validate_time_slot_settings(settings)
    meetings, length_minutes = module_meetings_per_week(module, increment)
    day_patterns = DAY_PATTERNS.get(meetings, DAY_PATTERNS[1])

    options: list[dict[str, str | int]] = []
    for days_index, days in enumerate(day_patterns):
        valid_starts = range(start_minutes, end_minutes - length_minutes + 1, increment)
        for start_index, start_minute in enumerate(valid_starts):
            options.append(
                {
                    "days": days,
                    "start": minutes_to_slot(start_minute),
                    "length": length_minutes // SLOT_MINUTES,
                    "weeks": weeks,
                    # Earlier, regular weekday slots are preferred. Later slots remain valid.
                    "penalty": days_index + start_index,
                }
            )
    if not options:
        raise ValueError(
            f"Module {module.code} has no valid time options. "
            f"Check the teaching window ({settings.start_time}-{settings.end_time}) "
            f"and duration increment ({settings.duration_increment_minutes} minutes)."
        )
    return options


def make_weeks_bit(nr_weeks: int, one_based_week: int) -> str:
    """Create a weeks bitstring with exactly one selected week."""

    return "".join("1" if idx == one_based_week else "0" for idx in range(1, nr_weeks + 1))


def make_days_bit(weekday: int) -> str:
    """Create a days bitstring with exactly one selected day."""

    return "".join("1" if idx == weekday else "0" for idx in range(DEFAULT_DAY_COUNT))


def add_xml_comment(parent: ET.Element, text: str) -> None:
    """Append a formatted XML comment for human-readable conversion notes."""

    parent.append(ET.Comment(f" {text} "))


def build_xml(
    rooms: list[RoomRow],
    modules: list[ModuleRow],
    cohorts: list[CohortRow],
    cohort_modules: dict[int, tuple[ModuleRow, ...]],
    nr_weeks: int,
    regular_weekdays_by_week: dict[int, set[int]],
    unavailable_days: list[tuple[int, int]],
    cohort_size: int,
    time_settings: TimeSlotSettings,
) -> ET.ElementTree:
    """Build an ITC 2019-compatible XML document."""

    root = ET.Element(
        "problem",
        {
            "name": "AITTO-Dataset-2025-26",
            "nrDays": str(DEFAULT_DAY_COUNT),
            "nrWeeks": str(nr_weeks),
            "slotsPerDay": str(SLOTS_PER_DAY),
        },
    )
    add_xml_comment(
        root,
        "Generated from dataset/*.xlsx by scripts/convert_dataset_to_itc2019.py. "
        "Cohorts are expanded into synthetic students because the source does not contain individual enrollment rows.",
    )

    ET.SubElement(
        root,
        "optimization",
        {
            "time": "1",
            "room": "1",
            "distribution": "1",
            "student": "1",
        },
    )

    rooms_el = ET.SubElement(root, "rooms")
    for room in rooms:
        room_el = ET.SubElement(
            rooms_el,
            "room",
            {
                "id": str(room.id),
                "capacity": str(room.capacity),
                "name": room.name,
                "type": room.room_type,
            },
        )
        for week, weekday in unavailable_days:
            ET.SubElement(
                room_el,
                "unavailable",
                {
                    "days": make_days_bit(weekday),
                    "start": "0",
                    "length": str(SLOTS_PER_DAY),
                    "weeks": make_weeks_bit(nr_weeks, week),
                },
            )

    courses_el = ET.SubElement(root, "courses")
    class_id = 1
    subpart_id = 1
    config_id = 1
    class_ids_by_module: dict[int, list[int]] = {}
    assigned_cohorts_by_module: dict[int, list[CohortRow]] = {
        module.id: [
            cohort
            for cohort in cohorts
            if any(selected.id == module.id for selected in cohort_modules.get(cohort.id, ()))
        ]
        for module in modules
    }

    for module in modules:
        course_el = ET.SubElement(
            courses_el,
            "course",
            {"id": str(module.id), "code": module.code, "title": module.title},
        )
        config_el = ET.SubElement(course_el, "config", {"id": str(config_id)})
        config_id += 1
        subpart_el = ET.SubElement(
            config_el,
            "subpart",
            {"id": str(subpart_id), "name": "Main"},
        )
        subpart_id += 1

        weeks = semester_weeks(regular_weekdays_by_week, module.semesters)
        time_options = make_time_options(module, weeks, time_settings)
        matching_rooms = [room for room in rooms if room.room_type.lower() == module.room_type.lower()]
        if not matching_rooms:
            matching_rooms = rooms

        class_ids_by_module[module.id] = []
        for cohort in assigned_cohorts_by_module.get(module.id, []):
            class_el = ET.SubElement(
                subpart_el,
                "class",
                {
                    "id": str(class_id),
                    "limit": str(cohort_size),
                    "name": f"{module.code}-{cohort.course}-{cohort.class_name}",
                    "cohort": f"{cohort.course}-{cohort.class_name}",
                },
            )
            class_ids_by_module[module.id].append(class_id)
            class_id += 1

            for room in matching_rooms:
                # Matching room types are preferred. The solver can still choose among them by capacity.
                penalty = 0 if room.capacity >= cohort_size else cohort_size - room.capacity
                ET.SubElement(class_el, "room", {"id": str(room.id), "penalty": str(penalty)})

            for option in time_options:
                ET.SubElement(
                    class_el,
                    "time",
                    {
                        "days": str(option["days"]),
                        "start": str(option["start"]),
                        "length": str(option["length"]),
                        "weeks": str(option["weeks"]),
                        "penalty": str(option["penalty"]),
                    },
                )

    distributions_el = ET.SubElement(root, "distributions")
    add_xml_comment(
        distributions_el,
        "Generated soft NotOverlap constraints keep sections of the same module from being scheduled at identical times where possible.",
    )
    for module in modules:
        ids = class_ids_by_module.get(module.id, [])
        if len(ids) < 2:
            continue
        distribution_el = ET.SubElement(distributions_el, "distribution", {"type": "NotOverlap", "penalty": "1"})
        for cid in ids:
            ET.SubElement(distribution_el, "class", {"id": str(cid)})

    students_el = ET.SubElement(root, "students")
    student_id = 1
    for cohort in cohorts:
        for offset in range(cohort_size):
            student_el = ET.SubElement(
                students_el,
                "student",
                {
                    "id": str(student_id),
                    "externalId": f"{cohort.course}-{cohort.class_name}-{offset + 1:02d}",
                },
            )
            for module in cohort_modules.get(cohort.id, ()):
                ET.SubElement(student_el, "course", {"id": str(module.id)})
            student_id += 1

    return ET.ElementTree(root)


def write_pretty_xml(tree: ET.ElementTree, path: Path) -> None:
    """Write XML with stable indentation for review and version control."""

    rough = ET.tostring(tree.getroot(), encoding="utf-8")
    parsed = minidom.parseString(rough)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(parsed.toprettyxml(indent="  ", encoding="utf-8").decode("utf-8"), encoding="utf-8")


def cohort_key(cohort: CohortRow) -> str:
    """Stable text key used by CLI mappings and the GUI."""

    return f"{cohort.course}-{cohort.class_name}"


def build_cohort_modules(
    cohorts: list[CohortRow],
    module_sets: dict[str, ModuleSet],
    cohort_module_labels: dict[str, str] | None = None,
) -> dict[int, tuple[ModuleRow, ...]]:
    """
    Resolve cohort -> module-set selections.

    If no mapping is provided for a cohort, ``None`` is used (no modules assigned).
    """

    if not module_sets:
        raise ValueError("At least one module set workbook is required.")

    labels = list(module_sets)
    cohort_module_labels = cohort_module_labels or {}

    resolved: dict[int, tuple[ModuleRow, ...]] = {}
    for cohort in cohorts:
        selected_label = cohort_module_labels.get(cohort_key(cohort), NONE_MODULE_SET)
        if selected_label == NONE_MODULE_SET:
            resolved[cohort.id] = ()
            continue
        if selected_label not in module_sets:
            known = ", ".join([NONE_MODULE_SET, *labels])
            raise ValueError(f"Unknown module set '{selected_label}' for cohort {cohort_key(cohort)}. Known: {known}")
        resolved[cohort.id] = module_sets[selected_label].modules
    return resolved


def parse_cohort_module_assignments(assignments: Iterable[str]) -> dict[str, str]:
    """Parse CLI values like ``FS123456-1A=modules_set_A``."""

    parsed: dict[str, str] = {}
    for value in assignments:
        if "=" not in value:
            raise ValueError(f"Invalid cohort module assignment '{value}'. Use COHORT-CLASS=MODULE_SET_LABEL.")
        cohort, label = value.split("=", 1)
        parsed[cohort.strip()] = label.strip()
    return parsed


def convert_dataset(
    dataset_dir: Path,
    output: Path,
    cohort_size: int,
    classroom_file: Path | None = None,
    students_file: Path | None = None,
    calendar_file: Path | None = None,
    module_files: Iterable[Path] | None = None,
    cohort_module_labels: dict[str, str] | None = None,
    time_settings: TimeSlotSettings | None = None,
) -> ConversionStats:
    """Convert selected source files into an ITC 2019-style XML file."""

    classroom_file = classroom_file or dataset_dir / "classroom.xlsx"
    students_file = students_file or dataset_dir / "students.xlsx"
    calendar_file = calendar_file or dataset_dir / "academic_calendar_2025_26.xlsx"
    module_file_list = list(module_files or discover_module_files(dataset_dir))
    time_settings = time_settings or TimeSlotSettings()
    validate_time_slot_settings(time_settings)

    rooms = load_rooms(classroom_file)
    modules, module_sets = load_module_sets(module_file_list)
    cohorts = load_cohorts(students_file)
    cohort_modules = build_cohort_modules(cohorts, module_sets, cohort_module_labels)
    nr_weeks, regular_weekdays_by_week, unavailable_days = load_calendar(calendar_file)

    active_module_ids = {
        module.id
        for selected_modules in cohort_modules.values()
        for module in selected_modules
    }
    # Cohorts assigned to "None" have no modules and are intentionally omitted
    # from the XML so the solver does not create unnecessary student variables.
    active_cohorts = [cohort for cohort in cohorts if cohort_modules.get(cohort.id)]
    active_modules = [module for module in modules if module.id in active_module_ids]
    class_count = sum(len(selected_modules) for selected_modules in cohort_modules.values())

    tree = build_xml(
        rooms=rooms,
        modules=active_modules,
        cohorts=active_cohorts,
        cohort_modules=cohort_modules,
        nr_weeks=nr_weeks,
        regular_weekdays_by_week=regular_weekdays_by_week,
        unavailable_days=unavailable_days,
        cohort_size=cohort_size,
        time_settings=time_settings,
    )
    write_pretty_xml(tree, output)

    return ConversionStats(
        output=output,
        rooms=len(rooms),
        modules=len(active_modules),
        cohorts=len(active_cohorts),
        classes=class_count,
        synthetic_students=len(active_cohorts) * cohort_size,
    )


def parse_args() -> argparse.Namespace:
    """Parse PowerShell-friendly command line arguments."""

    parser = argparse.ArgumentParser(description="Convert local Excel dataset to ITC 2019 XML.")
    parser.add_argument("--dataset-dir", type=Path, default=DATASET_DIR, help="Folder containing source .xlsx files.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output ITC 2019 XML path.")
    parser.add_argument("--classroom-file", type=Path, help="Classroom workbook path.")
    parser.add_argument("--students-file", type=Path, help="Cohort/student workbook path.")
    parser.add_argument("--calendar-file", type=Path, help="Academic calendar workbook path.")
    parser.add_argument(
        "--module-file",
        type=Path,
        action="append",
        dest="module_files",
        help="Module-set workbook path. Can be provided multiple times.",
    )
    parser.add_argument(
        "--cohort-module",
        action="append",
        default=[],
        help="Assign a module set to a cohort, e.g. FS123456-1A=modules_set_A. Can be repeated.",
    )
    parser.add_argument(
        "--cohort-size",
        type=int,
        default=DEFAULT_COHORT_SIZE,
        help="Synthetic student count per cohort row in students.xlsx.",
    )
    parser.add_argument(
        "--teaching-start",
        default=DEFAULT_TEACHING_START_TIME,
        help="Earliest generated teaching start time in HH:MM format.",
    )
    parser.add_argument(
        "--teaching-end",
        default=DEFAULT_TEACHING_END_TIME,
        help="Latest generated teaching end time in HH:MM format.",
    )
    parser.add_argument(
        "--duration-increment-minutes",
        type=int,
        default=DEFAULT_DURATION_INCREMENT_MINUTES,
        help="Round module durations and generated starts to this minute increment; must be a multiple of 30.",
    )
    parser.add_argument("--gui", action="store_true", help="Open the Tkinter converter GUI.")
    return parser.parse_args()


class ConverterGui:
    """Small Tkinter UI for selecting source workbooks and cohort module sets."""

    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import ttk

        self.tk = tk
        self.ttk = ttk
        self.root = tk.Tk()
        self.root.title("aiTTO ITC 2019 Dataset Converter")
        self.root.geometry("1180x760")

        self.dataset_dir = tk.StringVar(value=str(DATASET_DIR))
        self.classroom_file = tk.StringVar(value=str(DATASET_DIR / "classroom.xlsx"))
        self.students_file = tk.StringVar(value=str(DATASET_DIR / "students.xlsx"))
        self.calendar_file = tk.StringVar(value=str(DATASET_DIR / "academic_calendar_2025_26.xlsx"))
        self.output_file = tk.StringVar(value=str(DEFAULT_OUTPUT))
        self.cohort_size = tk.StringVar(value=str(DEFAULT_COHORT_SIZE))
        self.teaching_start = tk.StringVar(value=DEFAULT_TEACHING_START_TIME)
        self.teaching_end = tk.StringVar(value=DEFAULT_TEACHING_END_TIME)
        self.duration_increment = tk.StringVar(value=str(DEFAULT_DURATION_INCREMENT_MINUTES))
        self.selected_module_set = tk.StringVar()

        self.module_files: list[Path] = discover_module_files(DATASET_DIR)
        self.cohorts: list[CohortRow] = []
        self.module_sets: dict[str, ModuleSet] = {}
        self.cohort_module_labels: dict[str, str] = {}

        self._build()
        self.reload_sources()

    def _build(self) -> None:
        """Create all GUI widgets."""

        tk = self.tk
        ttk = self.ttk

        outer = ttk.Frame(self.root, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        files = ttk.LabelFrame(outer, text="Dataset files", padding=10)
        files.pack(fill=tk.X)
        self._path_row(files, "Dataset folder", self.dataset_dir, self.browse_dataset_dir, 0)
        self._path_row(files, "Classroom workbook", self.classroom_file, lambda: self.browse_file(self.classroom_file), 1)
        self._path_row(files, "Students/cohorts workbook", self.students_file, lambda: self.browse_file(self.students_file), 2)
        self._path_row(files, "Academic calendar workbook", self.calendar_file, lambda: self.browse_file(self.calendar_file), 3)
        self._path_row(files, "Output XML", self.output_file, self.browse_output_file, 4)

        options = ttk.Frame(outer)
        options.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(options, text="Synthetic students per cohort").pack(side=tk.LEFT)
        ttk.Entry(options, textvariable=self.cohort_size, width=8).pack(side=tk.LEFT, padx=(8, 16))
        ttk.Label(options, text="Teaching start").pack(side=tk.LEFT)
        ttk.Entry(options, textvariable=self.teaching_start, width=7).pack(side=tk.LEFT, padx=(8, 16))
        ttk.Label(options, text="Teaching end").pack(side=tk.LEFT)
        ttk.Entry(options, textvariable=self.teaching_end, width=7).pack(side=tk.LEFT, padx=(8, 16))
        ttk.Label(options, text="Duration step (mins)").pack(side=tk.LEFT)
        ttk.Entry(options, textvariable=self.duration_increment, width=6).pack(side=tk.LEFT, padx=(8, 16))
        ttk.Button(options, text="Reload dataset", command=self.reload_sources).pack(side=tk.LEFT)
        ttk.Button(options, text="Generate ITC XML", command=self.generate).pack(side=tk.RIGHT)
        ttk.Button(options, text="View Output XML", command=self.open_xml_viewer).pack(side=tk.RIGHT, padx=(0, 8))

        module_frame = ttk.LabelFrame(outer, text="Module-set workbooks", padding=10)
        module_frame.pack(fill=tk.X, pady=(10, 0))
        self.module_list = tk.Listbox(module_frame, height=4)
        self.module_list.pack(side=tk.LEFT, fill=tk.X, expand=True)
        module_buttons = ttk.Frame(module_frame)
        module_buttons.pack(side=tk.LEFT, padx=(10, 0), fill=tk.Y)
        ttk.Button(module_buttons, text="Add", command=self.add_module_file).pack(fill=tk.X)
        ttk.Button(module_buttons, text="Remove", command=self.remove_module_file).pack(fill=tk.X, pady=(6, 0))

        assign_frame = ttk.LabelFrame(outer, text="Cohort module-set assignment", padding=10)
        assign_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        left = ttk.Frame(assign_frame)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        columns = ("cohort", "module_set")
        self.cohort_tree = ttk.Treeview(left, columns=columns, show="headings", height=12)
        self.cohort_tree.heading("cohort", text="Cohort")
        self.cohort_tree.heading("module_set", text="Assigned module set")
        self.cohort_tree.column("cohort", width=280)
        self.cohort_tree.column("module_set", width=260)
        self.cohort_tree.pack(fill=tk.BOTH, expand=True)

        right = ttk.Frame(assign_frame)
        right.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 0))
        ttk.Label(right, text="Module set for selected cohort(s)").pack(anchor=tk.W)
        self.module_combo = ttk.Combobox(right, textvariable=self.selected_module_set, state="readonly", width=28)
        self.module_combo.pack(fill=tk.X, pady=(4, 8))
        ttk.Button(right, text="Apply to selected", command=self.apply_module_set_to_selected).pack(fill=tk.X)
        ttk.Button(right, text="Apply to all cohorts", command=self.apply_module_set_to_all).pack(fill=tk.X, pady=(6, 0))

        log_frame = ttk.LabelFrame(outer, text="Log", padding=10)
        log_frame.pack(fill=tk.BOTH, pady=(10, 0))
        self.log_text = tk.Text(log_frame, height=8, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _path_row(self, parent, label: str, variable, command, row: int) -> None:
        """Add a label, entry, and browse button for a path option."""

        ttk = self.ttk
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=2)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=8, pady=2)
        ttk.Button(parent, text="Browse", command=command).grid(row=row, column=2, pady=2)
        parent.columnconfigure(1, weight=1)

    def browse_dataset_dir(self) -> None:
        """Select the dataset folder and auto-fill standard workbook paths."""

        from tkinter import filedialog

        selected = filedialog.askdirectory(initialdir=self.dataset_dir.get(), title="Select dataset folder")
        if not selected:
            return
        dataset_dir = Path(selected)
        self.dataset_dir.set(str(dataset_dir))
        self.classroom_file.set(str(dataset_dir / "classroom.xlsx"))
        self.students_file.set(str(dataset_dir / "students.xlsx"))
        self.calendar_file.set(str(dataset_dir / "academic_calendar_2025_26.xlsx"))
        self.output_file.set(str(dataset_dir / "itc2019" / "aitto_dataset.xml"))
        self.module_files = discover_module_files(dataset_dir)
        self.reload_sources()

    def browse_file(self, variable) -> None:
        """Browse for an Excel source file."""

        from tkinter import filedialog

        selected = filedialog.askopenfilename(
            initialdir=self.dataset_dir.get(),
            filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")],
        )
        if selected:
            variable.set(selected)

    def browse_output_file(self) -> None:
        """Browse for the output XML location."""

        from tkinter import filedialog

        selected = filedialog.asksaveasfilename(
            initialdir=str(Path(self.output_file.get()).parent),
            defaultextension=".xml",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")],
        )
        if selected:
            self.output_file.set(selected)

    def add_module_file(self) -> None:
        """Add another module-set workbook."""

        from tkinter import filedialog

        selected = filedialog.askopenfilename(
            initialdir=self.dataset_dir.get(),
            filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")],
        )
        if selected:
            path = Path(selected)
            if path not in self.module_files:
                self.module_files.append(path)
            self.reload_sources()

    def remove_module_file(self) -> None:
        """Remove selected module-set workbooks."""

        selected_indexes = list(self.module_list.curselection())
        for index in reversed(selected_indexes):
            del self.module_files[index]
        self.reload_sources()

    def reload_sources(self) -> None:
        """Reload cohorts and module sets from the selected source files."""

        try:
            self.cohorts = load_cohorts(Path(self.students_file.get()))
            _, self.module_sets = load_module_sets(self.module_files)
        except Exception as exc:  # noqa: BLE001 - surfaced to a user-facing log.
            self.log(f"Could not reload sources: {exc}")
            return

        labels = list(self.module_sets)
        combo_values = [NONE_MODULE_SET, *labels]
        self.selected_module_set.set(NONE_MODULE_SET)
        self.module_combo["values"] = combo_values

        self.module_list.delete(0, self.tk.END)
        for path in self.module_files:
            label = path.stem
            count = len(self.module_sets[label].modules) if label in self.module_sets else 0
            self.module_list.insert(self.tk.END, f"{label} ({count} modules) - {path}")

        for item in self.cohort_tree.get_children():
            self.cohort_tree.delete(item)
        for cohort in self.cohorts:
            key = cohort_key(cohort)
            self.cohort_module_labels.setdefault(key, NONE_MODULE_SET)
            stored_label = self.cohort_module_labels[key]
            if stored_label != NONE_MODULE_SET and stored_label not in self.module_sets:
                stored_label = NONE_MODULE_SET
                self.cohort_module_labels[key] = stored_label
            self.cohort_tree.insert("", self.tk.END, iid=key, values=(key, self.cohort_module_labels[key]))

        self.log(f"Loaded {len(self.cohorts)} cohorts and {len(labels)} module set(s).")

    def apply_module_set_to_selected(self) -> None:
        """Apply the selected module set to highlighted cohorts."""

        label = self.selected_module_set.get()
        if not label:
            return
        for item in self.cohort_tree.selection():
            self.cohort_module_labels[item] = label
            self.cohort_tree.item(item, values=(item, label))

    def apply_module_set_to_all(self) -> None:
        """Apply the selected module set to every loaded cohort."""

        label = self.selected_module_set.get()
        if not label:
            return
        for cohort in self.cohorts:
            key = cohort_key(cohort)
            self.cohort_module_labels[key] = label
            self.cohort_tree.item(key, values=(key, label))

    def generate(self) -> None:
        """Run conversion using the GUI selections."""

        try:
            stats = convert_dataset(
                dataset_dir=Path(self.dataset_dir.get()),
                output=Path(self.output_file.get()),
                cohort_size=safe_int(self.cohort_size.get(), DEFAULT_COHORT_SIZE),
                classroom_file=Path(self.classroom_file.get()),
                students_file=Path(self.students_file.get()),
                calendar_file=Path(self.calendar_file.get()),
                module_files=self.module_files,
                cohort_module_labels=self.cohort_module_labels,
                time_settings=TimeSlotSettings(
                    start_time=self.teaching_start.get(),
                    end_time=self.teaching_end.get(),
                    duration_increment_minutes=safe_int(
                        self.duration_increment.get(),
                        DEFAULT_DURATION_INCREMENT_MINUTES,
                    ),
                ),
            )
        except Exception as exc:  # noqa: BLE001 - surfaced to a user-facing dialog.
            from tkinter import messagebox

            messagebox.showerror("Conversion failed", str(exc))
            self.log(f"Conversion failed: {exc}")
            return

        self.log(format_stats(stats))
        from tkinter import messagebox

        messagebox.showinfo("Conversion complete", f"Wrote {stats.output}")
        self.open_xml_viewer(Path(stats.output))

    def open_xml_viewer(self, xml_path: Path | None = None) -> None:
        """Open a separate window to inspect the generated ITC XML instance."""

        from tkinter import messagebox

        path = xml_path or Path(self.output_file.get())
        if not path.exists():
            messagebox.showwarning("XML not found", f"Cannot find XML file:\n{path}")
            return

        try:
            root = ET.parse(path).getroot()
            raw_xml = path.read_text(encoding="utf-8")
        except Exception as exc:  # noqa: BLE001 - surfaced to a user-facing dialog.
            messagebox.showerror("Could not open XML", str(exc))
            return

        tk = self.tk
        ttk = self.ttk
        window = tk.Toplevel(self.root)
        window.title(f"ITC 2019 XML Viewer - {path.name}")
        window.geometry("1120x760")

        notebook = ttk.Notebook(window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        summary_tab = ttk.Frame(notebook, padding=10)
        tree_tab = ttk.Frame(notebook, padding=10)
        raw_tab = ttk.Frame(notebook, padding=10)
        notebook.add(summary_tab, text="Summary")
        notebook.add(tree_tab, text="Structure")
        notebook.add(raw_tab, text="Raw XML")

        summary = tk.Text(summary_tab, wrap=tk.WORD)
        summary.pack(fill=tk.BOTH, expand=True)
        summary.insert(tk.END, self.xml_summary_text(root, path))
        summary.configure(state=tk.DISABLED)

        tree_frame = ttk.Frame(tree_tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(tree_frame, columns=("details",), show="tree headings")
        tree.heading("#0", text="Item")
        tree.heading("details", text="Details")
        tree.column("#0", width=420)
        tree.column("details", width=600)
        y_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        self.populate_xml_structure_tree(tree, root)

        raw = tk.Text(raw_tab, wrap=tk.NONE)
        raw_y = ttk.Scrollbar(raw_tab, orient=tk.VERTICAL, command=raw.yview)
        raw_x = ttk.Scrollbar(raw_tab, orient=tk.HORIZONTAL, command=raw.xview)
        raw.configure(yscrollcommand=raw_y.set, xscrollcommand=raw_x.set)
        raw.grid(row=0, column=0, sticky="nsew")
        raw_y.grid(row=0, column=1, sticky="ns")
        raw_x.grid(row=1, column=0, sticky="ew")
        raw_tab.rowconfigure(0, weight=1)
        raw_tab.columnconfigure(0, weight=1)
        raw.insert(tk.END, raw_xml)
        raw.configure(state=tk.DISABLED)

    def xml_summary_text(self, root: ET.Element, path: Path) -> str:
        """Build a readable summary of an ITC 2019 XML file."""

        rooms = root.findall("./rooms/room")
        courses = root.findall("./courses/course")
        classes = root.findall("./courses/course/config/subpart/class")
        distributions = root.findall("./distributions/distribution")
        students = root.findall("./students/student")
        time_options = sum(len(the_class.findall("./time")) for the_class in classes)
        room_options = sum(len(the_class.findall("./room")) for the_class in classes)
        unavailable = sum(len(room.findall("./unavailable")) for room in rooms)
        student_course_requests = sum(len(student.findall("./course")) for student in students)

        cohorts = sorted(
            {
                str(the_class.get("cohort"))
                for the_class in classes
                if the_class.get("cohort")
            }
        )
        distribution_types: dict[str, int] = {}
        for distribution in distributions:
            distribution_type = distribution.get("type", "Unknown")
            distribution_types[distribution_type] = distribution_types.get(distribution_type, 0) + 1

        lines = [
            f"File: {path}",
            "",
            "Problem",
            f"  Name: {root.get('name', '')}",
            f"  Days per week: {root.get('nrDays', '')}",
            f"  Weeks: {root.get('nrWeeks', '')}",
            f"  Slots per day: {root.get('slotsPerDay', '')}",
            "",
            "Counts",
            f"  Rooms: {len(rooms)}",
            f"  Courses/modules: {len(courses)}",
            f"  Class sections: {len(classes)}",
            f"  Students: {len(students)}",
            f"  Cohorts: {len(cohorts)}",
            f"  Distribution constraints: {len(distributions)}",
            "",
            "Generated domains",
            f"  Room options across all classes: {room_options}",
            f"  Time options across all classes: {time_options}",
            f"  Room unavailable entries: {unavailable}",
            f"  Student course requests: {student_course_requests}",
            "",
            "Optimization weights",
        ]

        optimization = root.find("./optimization")
        if optimization is not None:
            for key in ("time", "room", "distribution", "student"):
                lines.append(f"  {key}: {optimization.get(key, '0')}")

        lines.extend(["", "Cohorts"])
        lines.extend(f"  {cohort}" for cohort in cohorts)

        lines.extend(["", "Distribution constraint types"])
        if distribution_types:
            lines.extend(f"  {name}: {count}" for name, count in sorted(distribution_types.items()))
        else:
            lines.append("  None")

        return "\n".join(lines)

    def populate_xml_structure_tree(self, tree, root: ET.Element) -> None:
        """Populate the viewer tree with a concise ITC XML structure."""

        def insert(parent: str, text: str, details: str = "") -> str:
            return tree.insert(parent, self.tk.END, text=text, values=(details,))

        root_item = insert(
            "",
            f"Problem: {root.get('name', '')}",
            f"days={root.get('nrDays')} weeks={root.get('nrWeeks')} slotsPerDay={root.get('slotsPerDay')}",
        )

        optimization = root.find("./optimization")
        if optimization is not None:
            insert(
                root_item,
                "Optimization weights",
                ", ".join(f"{key}={value}" for key, value in optimization.attrib.items()),
            )

        rooms_item = insert(root_item, "Rooms", f"{len(root.findall('./rooms/room'))} rooms")
        for room in root.findall("./rooms/room"):
            insert(
                rooms_item,
                f"Room {room.get('id')} - {room.get('name', '')}",
                (
                    f"type={room.get('type', '')}, capacity={room.get('capacity', '')}, "
                    f"unavailable={len(room.findall('./unavailable'))}"
                ),
            )

        courses_item = insert(root_item, "Courses", f"{len(root.findall('./courses/course'))} courses")
        for course in root.findall("./courses/course"):
            course_item = insert(
                courses_item,
                f"Course {course.get('id')} - {course.get('code', '')}",
                course.get("title", ""),
            )
            for config in course.findall("./config"):
                config_item = insert(course_item, f"Config {config.get('id')}")
                for subpart in config.findall("./subpart"):
                    subpart_item = insert(
                        config_item,
                        f"Subpart {subpart.get('id')} - {subpart.get('name', '')}",
                        f"{len(subpart.findall('./class'))} classes",
                    )
                    for the_class in subpart.findall("./class"):
                        insert(
                            subpart_item,
                            f"Class {the_class.get('id')} - {the_class.get('name', '')}",
                            (
                                f"cohort={the_class.get('cohort', '')}, limit={the_class.get('limit', '')}, "
                                f"rooms={len(the_class.findall('./room'))}, times={len(the_class.findall('./time'))}"
                            ),
                        )

        distributions_item = insert(
            root_item,
            "Distributions",
            f"{len(root.findall('./distributions/distribution'))} constraints",
        )
        for index, distribution in enumerate(root.findall("./distributions/distribution"), start=1):
            insert(
                distributions_item,
                f"{index}. {distribution.get('type', 'Unknown')}",
                (
                    f"penalty={distribution.get('penalty', '0')}, "
                    f"required={distribution.get('required', 'false')}, "
                    f"classes={len(distribution.findall('./class'))}"
                ),
            )

        students_item = insert(root_item, "Students", f"{len(root.findall('./students/student'))} students")
        for student in root.findall("./students/student"):
            insert(
                students_item,
                f"Student {student.get('id')}",
                f"externalId={student.get('externalId', '')}, courses={len(student.findall('./course'))}",
            )

        tree.item(root_item, open=True)
        tree.item(courses_item, open=True)

    def log(self, message: str) -> None:
        """Append a message to the GUI log."""

        self.log_text.insert(self.tk.END, message.rstrip() + "\n")
        self.log_text.see(self.tk.END)

    def run(self) -> None:
        """Start the Tkinter event loop."""

        self.root.mainloop()


def format_stats(stats: ConversionStats) -> str:
    """Return the conversion summary used by CLI and GUI."""

    return "\n".join(
        [
            f"Wrote {stats.output}",
            f"Rooms: {stats.rooms}",
            f"Courses/modules: {stats.modules}",
            f"Cohorts: {stats.cohorts}",
            f"Classes/sections: {stats.classes}",
            f"Synthetic students: {stats.synthetic_students}",
        ]
    )


def main() -> None:
    """Load source workbooks and write the ITC 2019 XML instance."""

    args = parse_args()
    if args.gui:
        ConverterGui().run()
        return

    stats = convert_dataset(
        dataset_dir=args.dataset_dir,
        output=args.output,
        cohort_size=args.cohort_size,
        classroom_file=args.classroom_file,
        students_file=args.students_file,
        calendar_file=args.calendar_file,
        module_files=args.module_files,
        cohort_module_labels=parse_cohort_module_assignments(args.cohort_module),
        time_settings=TimeSlotSettings(
            start_time=args.teaching_start,
            end_time=args.teaching_end,
            duration_increment_minutes=args.duration_increment_minutes,
        ),
    )
    print(format_stats(stats))


if __name__ == "__main__":
    main()
