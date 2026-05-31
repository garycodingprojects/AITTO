"""
Build academic_calendar_2025_26.xlsx from the 2025/26 VTC-style academic calendar
(September 2025 through August 2026). Academic week numbers follow Sunday-based
rows anchored on 31 August 2025 (week 1).
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Output next to repo root
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "academic_calendar_2025_26.xlsx"

# Sunday that starts academic week 1 (matches printed calendar grids)
WEEK_ANCHOR = date(2025, 8, 31)

# Calendar span shown in the two source images
RANGE_START = date(2025, 9, 1)
RANGE_END = date(2026, 8, 31)

# Styling (ARGB without alpha prefix for openpyxl)
FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")  # semester commencement
FILL_GREEN = PatternFill("solid", fgColor="92D050")  # revision / exams
FILL_BLUE = PatternFill("solid", fgColor="BDD7EE")  # summer break
FONT_SUNDAY = Font(color="FF0000")
FONT_BOLD = Font(bold=True)
THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def academic_week_number(d: date) -> int:
    """Academic week index used in the printed calendar (1-based, Sun–Sat rows)."""
    return (d - WEEK_ANCHOR).days // 7 + 1


def daterange(start: date, end: date):
    """Inclusive date iterator."""
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def build_event_lookup() -> dict[date, tuple[str, str, str | None]]:
    """
    Map each special date to (event_type, remarks, highlight).
    highlight: 'yellow' | 'green' | 'blue' | None
    """
    events: dict[date, tuple[str, str, str | None]] = {}

    def add(
        day: date,
        event_type: str,
        remark: str,
        highlight: str | None = None,
    ) -> None:
        events[day] = (event_type, remark, highlight)

    def add_range(
        start: date,
        end: date,
        event_type: str,
        remark: str,
        highlight: str | None = None,
    ) -> None:
        for d in daterange(start, end):
            add(d, event_type, remark, highlight)

    # --- Image 1: Sep 2025 – Feb 2026 ---
    add(
        date(2025, 9, 1),
        "Semester Commencement",
        "Semester 1 Commencement",
        "yellow",
    )
    add(
        date(2025, 10, 1),
        "General Holiday",
        "General Holiday (National Day)",
    )
    add(
        date(2025, 10, 7),
        "General Holiday",
        "General Holiday (The day following Chinese mid-Autumn Festival)",
    )
    add(
        date(2025, 10, 29),
        "General Holiday",
        "General Holiday (Chung Yeung Festival)",
    )
    add(
        date(2025, 12, 6),
        "Graduation Ceremony",
        "Graduation Ceremony",
    )
    add_range(
        date(2025, 12, 22),
        date(2025, 12, 24),
        "College Holiday",
        "College Holidays",
    )
    add_range(
        date(2025, 12, 25),
        date(2025, 12, 26),
        "General Holiday",
        "General Holiday (Christmas Day & the first weekday after Christmas Day)",
    )
    add(date(2025, 12, 27), "College Holiday", "College Holidays")
    add_range(
        date(2025, 12, 29),
        date(2025, 12, 31),
        "College Holiday",
        "College Holidays",
    )
    add(
        date(2026, 1, 1),
        "General Holiday",
        "General Holiday (The first day of January)",
    )
    add_range(
        date(2026, 1, 15),
        date(2026, 1, 20),
        "Revision Week",
        "Revision Week (including Exams)",
        "green",
    )
    add(
        date(2026, 1, 28),
        "Semester Commencement",
        "Semester 2 Commencement",
        "yellow",
    )
    add(date(2026, 2, 16), "College Holiday", "College Holidays")
    add_range(
        date(2026, 2, 17),
        date(2026, 2, 19),
        "General Holiday",
        "General Holidays (Lunar New Year, the second day & third day of Lunar New Year)",
    )
    add(date(2026, 2, 20), "College Holiday", "College Holidays")

    # --- Image 2: Mar 2026 – Aug 2026 ---
    add(date(2026, 4, 2), "College Holiday", "College Holidays")
    add_range(
        date(2026, 4, 3),
        date(2026, 4, 4),
        "General Holiday",
        "General Holiday (The day following Good Friday)",
    )
    add(
        date(2026, 4, 6),
        "General Holiday",
        "General Holidays (The day following Ching Ming Festival)",
    )
    add(
        date(2026, 4, 7),
        "General Holiday",
        "General Holidays (The day following Easter Monday)",
    )
    add_range(
        date(2026, 4, 8),
        date(2026, 4, 9),
        "College Holiday",
        "College Holidays",
    )
    add(
        date(2026, 5, 1),
        "General Holiday",
        "General Holiday (Labour Day)",
    )
    add(
        date(2026, 5, 25),
        "General Holiday",
        "General Holiday (The day following the Birthday of the Buddha)",
    )
    add(
        date(2026, 6, 19),
        "General Holiday",
        "General Holiday (Tuen Ng Festival)",
    )
    for d in (date(2026, 6, 26), date(2026, 6, 29), date(2026, 6, 30)):
        add(
            d,
            "Revision Week",
            "Revision Week (including Exams)",
            "green",
        )
    add(
        date(2026, 7, 1),
        "General Holiday",
        "General Holiday (HKSAR Establishment Day)",
    )
    summer_remark = (
        "Summer Break (with exchange programmes for nominated students)"
    )
    add_range(
        date(2026, 7, 2),
        date(2026, 8, 31),
        "Summer Break",
        summer_remark,
        "blue",
    )

    return events


def symbol_for_event(event_type: str) -> str:
    """Calendar-grid symbol from the printed legend."""
    if event_type == "General Holiday":
        return "'"
    if event_type == "College Holiday":
        return "#"
    return ""


def write_daily_sheet(wb: Workbook, events: dict) -> None:
    """One row per day with metadata and cell formatting."""
    ws = wb.active
    ws.title = "Daily Calendar"

    headers = [
        "Date",
        "Week",
        "Day of Week",
        "Month",
        "Event Type",
        "Remarks",
        "Symbol",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = FONT_BOLD
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for d in daterange(RANGE_START, RANGE_END):
        event_type, remark, highlight = events.get(d, ("Regular", "", None))
        if d.weekday() == 6 and event_type == "Regular":  # Sunday
            event_type = "Sunday"

        ws.append(
            [
                d,
                academic_week_number(d),
                d.strftime("%A"),
                d.strftime("%B %Y"),
                event_type,
                remark,
                symbol_for_event(event_type) if event_type not in ("Regular", "Sunday") else "",
            ]
        )

        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = BORDER_ALL
            if col == 1:
                cell.number_format = "yyyy-mm-dd"

        # Sunday dates in red (column C = day name; column A date value styling)
        if d.weekday() == 6:
            ws.cell(row=row_idx, column=1).font = FONT_SUNDAY
            ws.cell(row=row_idx, column=3).font = FONT_SUNDAY

        if highlight == "yellow":
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = FILL_YELLOW
        elif highlight == "green":
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = FILL_GREEN
        elif highlight == "blue":
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = FILL_BLUE

        row_idx += 1

    ws.freeze_panes = "A2"
    widths = [12, 6, 14, 16, 22, 70, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_grid_sheet(wb: Workbook, events: dict) -> None:
    """
    Recreate the printed Sun–Sat grid: one row per calendar week with
    week number, month label, seven day cells, and merged remarks.
    """
    ws = wb.create_sheet("Calendar Grid")
    headers = ["Week", "Month", "S", "M", "T", "W", "T", "F", "S", "Remarks"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = FONT_BOLD
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = BORDER_ALL
        c.alignment = Alignment(horizontal="center")

    # Collect remark lines grouped by month for the right column
    month_remarks: dict[tuple[int, int], list[str]] = {}

    def month_key(d: date) -> tuple[int, int]:
        return (d.year, d.month)

    remark_lines = [
        (date(2025, 9, 1), "1/9: Semester 1 Commencement"),
        (date(2025, 10, 1), "1/10: General Holiday (National Day)"),
        (date(2025, 10, 7), "7/10: General Holiday (The day following Chinese mid-Autumn Festival)"),
        (date(2025, 10, 29), "29/10: General Holiday (Chung Yeung Festival)"),
        (date(2025, 12, 6), "6/12: Graduation Ceremony"),
        (date(2025, 12, 22), "22-24/12: College Holidays"),
        (date(2025, 12, 25), "25-26/12: General Holiday (Christmas Day & the first weekday after Christmas Day)"),
        (date(2025, 12, 27), "27 & 29-31/12: College Holidays"),
        (date(2026, 1, 1), "1/1: General Holiday (The first day of January)"),
        (date(2026, 1, 15), "15-20/1: Revision Week (including Exams)"),
        (date(2026, 1, 28), "28/1: Semester 2 Commencement"),
        (date(2026, 2, 16), "16 & 20/2: College Holidays"),
        (date(2026, 2, 17), "17-19/2: General Holidays (Lunar New Year, the second day & third day of Lunar New Year)"),
        (date(2026, 4, 2), "2/4: College Holidays"),
        (date(2026, 4, 3), "2-4/4: General Holiday (The day following Good Friday)"),
        (date(2026, 4, 6), "6/4: General Holidays (The day following Ching Ming Festival)"),
        (date(2026, 4, 7), "7/4: General Holidays (The day following Easter Monday)"),
        (date(2026, 4, 8), "8-9/4: College Holidays"),
        (date(2026, 5, 1), "1/5: General Holiday (Labour Day)"),
        (date(2026, 5, 25), "25/5: General Holiday (The day following the Birthday of the Buddha)"),
        (date(2026, 6, 19), "19/6: General Holiday (Tuen Ng Festival)"),
        (date(2026, 6, 26), "26, 29-30/6: Revision Week (including Exams)"),
        (date(2026, 7, 1), "1/7: General Holiday (HKSAR Establishment Day)"),
        (date(2026, 7, 2), "2/7 – 31/8: Summer Break (with exchange programmes for nominated students)"),
    ]
    for d, text in remark_lines:
        month_remarks.setdefault(month_key(d), []).append(text)

    row = 2
    week_start = WEEK_ANCHOR
    last_month: tuple[int, int] | None = None
    month_row_start = 2

    while week_start <= RANGE_END:
        week_num = academic_week_number(week_start)
        days = [week_start + timedelta(days=i) for i in range(7)]

        # Month label on first row of each calendar month in column B
        ref_day = next((d for d in days if RANGE_START <= d <= RANGE_END), days[0])
        mk = month_key(ref_day)
        month_label = ""
        if mk != last_month and RANGE_START <= ref_day <= RANGE_END:
            month_label = ref_day.strftime("%B %Y")
            last_month = mk

        week_vals: list = [week_num, month_label]
        for d in days:
            if RANGE_START <= d <= RANGE_END:
                et, _, hl = events.get(d, ("Regular", "", None))
                sym = symbol_for_event(et) if et not in ("Regular", "Sunday") else ""
                display = f"{d.day}{sym}" if sym else str(d.day)
                week_vals.append(display)
            else:
                week_vals.append("")

        # Remarks: show month block text on first week row of that month
        remarks_cell = ""
        if month_label:
            remarks_cell = "\n".join(month_remarks.get(mk, []))
        week_vals.append(remarks_cell)

        ws.append(week_vals)
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER_ALL
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # Style individual day cells (cols 3–9)
        for i, d in enumerate(days):
            if not (RANGE_START <= d <= RANGE_END):
                continue
            cell = ws.cell(row=row, column=3 + i)
            et, _, hl = events.get(d, ("Regular", "", None))
            if d.weekday() == 6:
                cell.font = FONT_SUNDAY
            if hl == "yellow":
                cell.fill = FILL_YELLOW
            elif hl == "green":
                cell.fill = FILL_GREEN
            elif hl == "blue":
                cell.fill = FILL_BLUE

        row += 1
        week_start += timedelta(days=7)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    for letter in "CDEFGHI":
        ws.column_dimensions[letter].width = 5
    ws.column_dimensions["J"].width = 55

    # Legend block below the grid
    legend_row = row + 2
    ws.cell(row=legend_row, column=1, value="Legend").font = FONT_BOLD
    legends = [
        ("Yellow fill", "Semester Commencement"),
        ("Red text / '", "General Holiday"),
        ("Black box (see daily sheet)", "VTC Graduation Ceremony"),
        ("Green fill", "Revision Week & Assessment"),
        ("Red #", "College Holiday"),
        ("Light blue fill", "Summer Break"),
    ]
    for i, (sym, desc) in enumerate(legends, start=1):
        ws.cell(row=legend_row + i, column=1, value=sym)
        ws.cell(row=legend_row + i, column=2, value=desc)


def write_key_dates_sheet(wb: Workbook, events: dict) -> None:
    """Condensed list of non-regular days only."""
    ws = wb.create_sheet("Key Dates")
    ws.append(["Date", "Week", "Event Type", "Remarks"])
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = FONT_BOLD

    r = 2
    for d in daterange(RANGE_START, RANGE_END):
        if d not in events:
            continue
        et, remark, _ = events[d]
        ws.append([d, academic_week_number(d), et, remark])
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        r += 1

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["D"].width = 70


def main() -> None:
    events = build_event_lookup()
    wb = Workbook()
    write_daily_sheet(wb, events)
    write_grid_sheet(wb, events)
    write_key_dates_sheet(wb, events)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
